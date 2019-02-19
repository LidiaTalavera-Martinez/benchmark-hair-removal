from openpyxl import Workbook, load_workbook
import os


def get_excel_file(excel_name, sheet_name):

    if os.path.isfile(excel_name):
        workbook = load_workbook(excel_name)

    else:
        workbook = Workbook()
    worksheet = workbook.create_sheet(sheet_name)

    return workbook, worksheet
