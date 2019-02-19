import numpy as np
import pandas as pd
from scipy.stats import shapiro, wilcoxon, wilcoxon, ttest_rel
from openpyxl import Workbook, load_workbook
import os
import statistics as stats
from xlsx_file import get_excel_file
import constants

def statistic_test(dfs,metric_name,test_excel_name,methods_used):

    workbook, worksheet = get_excel_file(test_excel_name, metric_name)

    for i in range(1, dfs.shape[0], methods_used+1):
        if i == 1:
            metric_values = dfs.loc[i:i + 5][metric_name].values
        else:
            metric_values = np.vstack((metric_values, dfs.loc[i:i + 5][metric_name].values))

    metric_values = np.transpose(metric_values)

    num_row = 1
    methods_labels = constants.methods

    for method1 in range(len(methods_labels)):
        for method2 in range(method1 + 1, len(methods_labels)):
            dif_data = abs(metric_values[method1, :] - metric_values[method2, :])

            # normality test
            stat, p = shapiro(dif_data)
            alpha = constants.alpha_thr
            if p > alpha:  # Sample looks Gaussian (fail to reject H0)
                T, p_value = ttest_rel(metric_values[method1, :], metric_values[method2, :])
                test = constants.T_Test

            else:  # Sample does not look Gaussian (reject H0)
                T, p_value = wilcoxon(metric_values[method1, :], metric_values[method2, :])
                p_value = p_value
                test = constants.W_Test

            """EXCEL"""
            worksheet.cell(row=num_row, column=1).value = (methods_labels[method1] + '-' + methods_labels[method2])
            worksheet.cell(row=num_row, column=2).value = p_value
            worksheet.cell(row=num_row, column=3).value = test
            worksheet.cell(row=num_row, column=4).value = stats.mean(metric_values[method1, :])
            worksheet.cell(row=num_row, column=5).value = stats.mean(metric_values[method2, :])
            if stats.mean(metric_values[method1, :]) >= stats.mean(metric_values[method2, :]):
                worksheet.cell(row=num_row, column=6).value = 'mean method1 >= mean method2'
            else:
                worksheet.cell(row=num_row, column=6).value = 'mean method1 < mean method2'

            num_row = num_row + 1
    workbook.save(test_excel_name)
    return
